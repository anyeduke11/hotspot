"""Phase 4 ExportService — 预生成静态 HTML + ETag 304 + 交互式筛选 + XLSX 导出。

- :func:`build_html`          实时构造「默认」HTML（无缓存, 慢路径, 兼容 ETag/304）
- :func:`get_or_build_html`   返回 (html, etag) — 优先返回缓存
- :func:`build_export_html`   实时构造「带筛选」的 light 风格 HTML（带筛选 UI + 预览表）
- :func:`build_xlsx`          实时构造 XLSX 文件 (3 列: 标题 / 来源 / 原文链接)
- 缓存文件位于 ``data/export_cache.html`` + ``data/export_cache.etag``
"""
from __future__ import annotations

import hashlib
import io
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from backend.config import BASE_DIR
from backend.domain.enums import Category, TimeRange
from backend.exceptions import InvalidParamException
from backend.logging_config import logger
from backend.repository.hotspot_repo import HotspotRepository

_hrepo = HotspotRepository()

# 缓存文件路径（位于 data/ 目录）
_DATA_DIR = BASE_DIR / "data"
_CACHE_HTML = _DATA_DIR / "export_cache.html"
_CACHE_ETAG = _DATA_DIR / "export_cache.etag"

# 类别色值 + 标签（与前端保持一致）
_CAT_LABELS: dict[str, str] = {
    "ai": "科技 / AI",
    "security": "网络安全",
    "finance": "金融 / 投资",
    "startup": "独立开发 / 创业",
    "bid": "招标资讯",
    "github": "GitHub 项目",
    "tech": "科技 / IT",
}
_CAT_COLORS: dict[str, str] = {
    "ai": "#00bcd4",
    "security": "#e85d5d",
    "finance": "#f0c929",
    "startup": "#7c6aff",
    "bid": "#e8891a",
    "github": "#7c6aff",
    "tech": "#00bcd4",
}

# ---------------------------------------------------------------------------
# 交互式导出: 筛选参数白名单
# ---------------------------------------------------------------------------
# 仅暴露给 /api/export 筛选 UI 的窗口, 避免 30d 之类的边角情况
_EXPORT_TIME_RANGES: tuple[str, ...] = ("24h", "3d", "7d")
_EXPORT_TYPE_TOKENS: tuple[str, ...] = ("hotspot", "bid")

# type token → 该 type 涵盖的 category 集合
_TYPE_TO_CATEGORIES: dict[str, tuple[Category, ...]] = {
    "hotspot": (
        Category.AI, Category.SECURITY, Category.FINANCE,
        Category.STARTUP, Category.TECH, Category.GITHUB,
    ),
    "bid": (Category.BID,),
}

# type token → 中文标签 (UI 显示)
_TYPE_LABELS: dict[str, str] = {
    "hotspot": "资讯",
    "bid": "标讯",
}

# type token → 简称 (紧凑显示, 用于复选框右侧)
_TYPE_LABELS_SHORT: dict[str, str] = {
    "hotspot": "资讯 (科技/AI / 安全 / 金融 / 创业 / GitHub)",
    "bid": "标讯 (招标)",
}

# Light 主题色值 (与 frontend/src/index.css [data-theme="light"] 保持一致)
_LIGHT = {
    "bg_primary": "#f4f4f8",
    "bg_card": "#ffffff",
    "bg_hover": "#eeeef4",
    "border": "#dcdce6",
    "text_primary": "#1a1a2e",
    "text_secondary": "#555570",
    "text_muted": "#9999aa",
    "accent_ai": "#00bcd4",
    "shadow": "0 1px 3px rgba(0, 0, 0, 0.04), 0 1px 2px rgba(0, 0, 0, 0.03)",
}

# 单次导出最大条数 (避免内存爆炸)
_EXPORT_MAX_ITEMS = 5000


def _esc(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# 筛选参数解析
# ---------------------------------------------------------------------------
def parse_export_filters(
    time_range: str = "7d",
    types: str = "hotspot,bid",
) -> Tuple[TimeRange, set[str]]:
    """解析 /api/export 筛选参数。

    Raises:
        InvalidParamException: 参数非法 (前端 fetch 会被 400 拦截, 用户看到错误提示)
    """
    if time_range not in _EXPORT_TIME_RANGES:
        raise InvalidParamException(
            f"time_range 必须是 {_EXPORT_TIME_RANGES} 之一, 收到: {time_range!r}"
        )
    # 解析 type tokens
    tokens = [t.strip() for t in (types or "").split(",") if t.strip()]
    if not tokens:
        raise InvalidParamException("types 不能为空 (至少选择 'hotspot' 或 'bid')")
    invalid = [t for t in tokens if t not in _EXPORT_TYPE_TOKENS]
    if invalid:
        raise InvalidParamException(
            f"types 含未知 token {invalid}, 允许值: {list(_EXPORT_TYPE_TOKENS)}"
        )
    return TimeRange(time_range), set(tokens)


def fetch_export_items(time_range: TimeRange, types: set[str]) -> list:
    """根据 type token 集合拉取数据, 按 ingested_at DESC 合并去重。"""
    if not types:
        return []
    # 展开 type → category 列表
    target_cats: list[Category] = []
    for t in types:
        target_cats.extend(_TYPE_TO_CATEGORIES[t])
    # 按 category 逐个 query (SQL 已合并 tech → ai)
    seen: set[str] = set()
    merged: list = []
    for cat in target_cats:
        batch, _ = _hrepo.query(
            category=cat,
            time_range=time_range,
            limit=_EXPORT_MAX_ITEMS,
        )
        for it in batch:
            if it.id in seen:
                continue
            seen.add(it.id)
            merged.append(it)
    # 全局按 ingested_at DESC 排序 (与首页一致)
    merged.sort(
        key=lambda x: x.ingested_at or x.published_at or "",
        reverse=True,
    )
    return merged[:_EXPORT_MAX_ITEMS]


# ---------------------------------------------------------------------------
# 实时构建 — 旧版兼容 (默认 7d + 全部 category)
# ---------------------------------------------------------------------------
def build_html() -> str:
    """构造完整的静态导出 HTML（无缓存, 默认 7d + 全部 type）。

    保留 ETag/304 流程: scheduler 调 :func:`rebuild_export_cache` 预热。
    """
    items, _ = _hrepo.query(category=None, time_range=TimeRange.D7, limit=200)

    # 类别计数
    cat_counts: dict[str, int] = {c.value: 0 for c in Category}
    for it in items:
        cat_counts[it.category.value] = cat_counts.get(it.category.value, 0) + 1

    # 卡片 HTML
    cards_html = ""
    for item in items:
        cat = item.category.value
        color = _CAT_COLORS.get(cat, "#888")
        label = _CAT_LABELS.get(cat, cat)
        title = _esc(item.title)
        summary = _esc(item.summary or "")
        source = _esc(item.source)
        url = _esc(str(item.url))
        cards_html += f"""
        <a href="{url}" target="_blank" rel="noopener" class="card" style="border-left:3px solid {color}">
            <div class="card-header">
                <span class="tag" style="background:{color}20;color:{color}">{label}</span>
            </div>
            <h3>{title}</h3>
            <p>{summary}</p>
            <div class="card-footer">
                <span>{source}</span>
                <span style="color:{color}">查看原文 &rarr;</span>
            </div>
        </a>"""

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>热点地图 - 静态导出</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:'JetBrains Mono','Fira Code',monospace; background:#0a0a0f; color:#e0e0e0; min-height:100vh; }}
.bg-grid {{ background-image:linear-gradient(rgba(0,188,212,0.03) 1px,transparent 1px),linear-gradient(90deg,rgba(0,188,212,0.03) 1px,transparent 1px); background-size:60px 60px; }}
.container {{ max-width:1280px; margin:0 auto; padding:24px 16px; }}
.header {{ display:flex; justify-content:space-between; align-items:center; border-bottom:1px solid #1e1e30; padding-bottom:16px; margin-bottom:24px; flex-wrap:wrap; gap:12px; }}
.header-left {{ display:flex; align-items:center; gap:12px; }}
.logo {{ width:40px; height:40px; border-radius:8px; background:linear-gradient(135deg,#00bcd4,#7c6aff); display:flex; align-items:center; justify-content:center; font-weight:bold; font-size:20px; }}
.header-info {{ color:#888899; font-size:12px; }}
.stats {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; margin-bottom:24px; }}
.stat-card {{ background:#12121a; border:1px solid #1e1e30; border-radius:12px; padding:16px; }}
.stat-title {{ font-size:11px; color:#888899; text-transform:uppercase; margin-bottom:8px; }}
.stat-value {{ font-size:24px; font-weight:bold; }}
.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(280px,1fr)); gap:16px; }}
.card {{ display:block; background:#12121a; border:1px solid #1e1e30; border-radius:12px; padding:20px; text-decoration:none; color:inherit; transition:all 0.3s; }}
.card:hover {{ transform:translateY(-2px); border-color:rgba(255,255,255,0.15); }}
.card-header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; }}
.tag {{ font-size:10px; font-weight:600; padding:2px 8px; border-radius:999px; }}
.card h3 {{ font-size:14px; line-height:1.5; margin-bottom:8px; color:#fff; }}
.card p {{ font-size:12px; color:#888899; line-height:1.5; margin-bottom:12px; display:-webkit-box; -webkit-line-clamp:2; -webkit-box-orient:vertical; overflow:hidden; }}
.card-footer {{ display:flex; justify-content:space-between; font-size:11px; color:#555; }}
.footer {{ text-align:center; padding:24px 0; border-top:1px solid #1e1e30; margin-top:32px; font-size:11px; color:#555; }}
@media (max-width:640px) {{ .grid {{ grid-template-columns:1fr; }} }}
</style>
</head>
<body class="bg-grid">
<div class="container">
    <div class="header">
        <div class="header-left">
            <div class="logo">H</div>
            <div>
                <h1 style="font-size:20px;">热点地图 <span style="font-size:12px;color:#888899;font-weight:normal;">HOTSPOT MAP</span></h1>
                <p style="font-size:12px;color:#888899;margin-top:2px;">{now} UTC 导出</p>
            </div>
        </div>
        <div class="header-info">共 {len(items)} 条热点</div>
    </div>

    <div class="stats">
        <div class="stat-card"><div class="stat-title">总计</div><div class="stat-value" style="color:#00c96a">{len(items)}</div></div>
        <div class="stat-card"><div class="stat-title">科技/AI</div><div class="stat-value" style="color:#00bcd4">{cat_counts.get('ai', 0)}</div></div>
        <div class="stat-card"><div class="stat-title">网络安全</div><div class="stat-value" style="color:#e85d5d">{cat_counts.get('security', 0)}</div></div>
        <div class="stat-card"><div class="stat-title">金融/投资</div><div class="stat-value" style="color:#f0c929">{cat_counts.get('finance', 0)}</div></div>
        <div class="stat-card"><div class="stat-title">独立开发/创业</div><div class="stat-value" style="color:#7c6aff">{cat_counts.get('startup', 0)}</div></div>
        <div class="stat-card"><div class="stat-title">招标资讯</div><div class="stat-value" style="color:#e8891a">{cat_counts.get('bid', 0)}</div></div>
    </div>

    <div class="grid">{cards_html}</div>

    <div class="footer">
        热点地图 &middot; 数据来源: aihot.virxact.com / AVD / THN / 新浪财经 / Hacker News<br>
        导出时间: {now} UTC
    </div>
</div>
</body>
</html>"""


# ---------------------------------------------------------------------------
# 实时构建 — 交互式导出 (light 风格, 带筛选 UI + 预览表)
# ---------------------------------------------------------------------------
def build_export_html(
    time_range: TimeRange,
    types: set[str],
    items: list,
) -> str:
    """构造「数据导出」页 HTML — light 风格, 与首页统一, 含筛选 UI + 预览表。

    筛选参数来自 URL query string, 用户修改后 form submit 触发 GET 请求重新渲染。
    """
    time_range_labels = {"24h": "24 小时", "3d": "3 天", "7d": "7 天"}
    tr_label = time_range_labels.get(time_range.value, time_range.value)

    # 类别计数 (仅在勾选的 type 范围内)
    cat_counts: dict[str, int] = {}
    for it in items:
        c = it.category.value
        cat_counts[c] = cat_counts.get(c, 0) + 1

    # 预览表 (最多前 100 条)
    preview_items = items[:100]
    rows_html = ""
    for idx, it in enumerate(preview_items, 1):
        title = _esc(it.title)
        source = _esc(it.source)
        url = _esc(str(it.url))
        cat_label = _CAT_LABELS.get(it.category.value, it.category.value)
        cat_color = _CAT_COLORS.get(it.category.value, "#888")
        rows_html += f"""
        <tr>
          <td class="col-num">{idx}</td>
          <td class="col-title"><a href="{url}" target="_blank" rel="noopener">{title}</a></td>
          <td class="col-src"><span class="tag" style="background:{cat_color}18;color:{cat_color}">{_esc(cat_label)}</span><br><span class="src-name">{source}</span></td>
          <td class="col-url"><a href="{url}" target="_blank" rel="noopener" class="url-link">{url}</a></td>
        </tr>"""

    # radio/checkbox 状态
    tr_checked = {tr: (tr == time_range.value) for tr in _EXPORT_TIME_RANGES}
    type_checked = {t: (t in types) for t in _EXPORT_TYPE_TOKENS}

    # 状态条: 当前筛选摘要
    types_label = " + ".join(_TYPE_LABELS[t] for t in types)
    summary = f"{tr_label} · {types_label} · 共 {len(items)} 条"
    full_count = len(items)
    preview_count = len(preview_items)
    overflow_note = ""
    if full_count > preview_count:
        overflow_note = f"（预览仅展示前 {preview_count} 条, 完整数据请下载 XLSX）"

    # 各 type 的数量 (用于复选框旁 tooltip)
    type_counts: dict[str, int] = {}
    for t in types:
        cats = _TYPE_TO_CATEGORIES[t]
        type_counts[t] = sum(cat_counts.get(c.value, 0) for c in cats)

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    L = _LIGHT
    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>热点地图 / 数据导出</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{
    font-family: 'JetBrains Mono', 'Fira Code', 'Cascadia Code', monospace;
    background: {L['bg_primary']};
    color: {L['text_primary']};
    min-height: 100vh;
    -webkit-font-smoothing: antialiased;
  }}
  .container {{ max-width: 1280px; margin: 0 auto; padding: 24px 16px 64px; }}
  /* ── Header ── */
  .header {{
    display: flex; align-items: center; justify-content: space-between;
    padding: 18px 22px; background: {L['bg_card']};
    border: 1px solid {L['border']}; border-radius: 14px;
    box-shadow: {L['shadow']}; margin-bottom: 18px;
  }}
  .header-left {{ display: flex; align-items: center; gap: 14px; }}
  .logo {{
    width: 42px; height: 42px; border-radius: 10px;
    background: linear-gradient(135deg, #00bcd4, #7c6aff);
    display: flex; align-items: center; justify-content: center;
    color: #fff; font-weight: 700; font-size: 20px;
  }}
  .header-title {{ font-size: 18px; font-weight: 600; letter-spacing: 0.02em; }}
  .header-subtitle {{ font-size: 11px; color: {L['text_muted']}; margin-top: 2px; }}
  .header-meta {{ font-size: 12px; color: {L['text_secondary']}; }}
  /* ── Filter card ── */
  .card {{
    background: {L['bg_card']}; border: 1px solid {L['border']};
    border-radius: 14px; box-shadow: {L['shadow']};
    padding: 22px; margin-bottom: 18px;
  }}
  .card-title {{
    font-size: 11px; text-transform: uppercase; letter-spacing: 0.08em;
    color: {L['text_muted']}; margin-bottom: 14px;
  }}
  .field-row {{
    display: flex; flex-wrap: wrap; gap: 18px 28px; align-items: center;
    margin-bottom: 14px;
  }}
  .field-row:last-of-type {{ margin-bottom: 0; }}
  .field-label {{
    font-size: 12px; color: {L['text_secondary']};
    min-width: 76px; letter-spacing: 0.04em;
  }}
  .opt {{ display: inline-flex; align-items: center; gap: 6px; cursor: pointer; }}
  .opt input {{ accent-color: {L['accent_ai']}; cursor: pointer; }}
  .opt-label {{ font-size: 13px; color: {L['text_primary']}; }}
  .opt-hint {{ font-size: 11px; color: {L['text_muted']}; margin-left: 2px; }}
  /* ── Actions row ── */
  .actions {{
    display: flex; align-items: center; gap: 12px;
    margin-top: 20px; padding-top: 18px;
    border-top: 1px dashed {L['border']};
  }}
  .btn {{
    display: inline-flex; align-items: center; gap: 6px;
    padding: 8px 18px; border-radius: 8px;
    font: inherit; font-size: 13px; font-weight: 600;
    cursor: pointer; transition: all 0.15s ease;
    border: 1px solid transparent;
  }}
  .btn-primary {{
    background: {L['accent_ai']}; color: #fff;
  }}
  .btn-primary:hover {{ filter: brightness(1.08); }}
  .btn-ghost {{
    background: transparent; color: {L['text_secondary']};
    border-color: {L['border']};
  }}
  .btn-ghost:hover {{ color: {L['text_primary']}; border-color: {L['text_muted']}; }}
  .actions-summary {{
    margin-left: auto; font-size: 12px; color: {L['text_secondary']};
  }}
  .actions-summary b {{ color: {L['text_primary']}; font-weight: 600; }}
  /* ── Preview table ── */
  .preview {{
    background: {L['bg_card']}; border: 1px solid {L['border']};
    border-radius: 14px; box-shadow: {L['shadow']};
    overflow: hidden;
  }}
  .preview-head {{
    padding: 14px 22px; border-bottom: 1px solid {L['border']};
    display: flex; align-items: baseline; justify-content: space-between;
  }}
  .preview-title {{ font-size: 13px; font-weight: 600; }}
  .preview-hint {{ font-size: 11px; color: {L['text_muted']}; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  th, td {{
    padding: 10px 14px; text-align: left;
    border-bottom: 1px solid {L['border']};
  }}
  th {{
    font-size: 10px; text-transform: uppercase; letter-spacing: 0.08em;
    color: {L['text_muted']}; font-weight: 600;
    background: {L['bg_hover']};
  }}
  td.col-num {{ color: {L['text_muted']}; width: 36px; text-align: right; }}
  td.col-title a {{ color: {L['text_primary']}; text-decoration: none; }}
  td.col-title a:hover {{ color: {L['accent_ai']}; text-decoration: underline; }}
  td.col-src {{ width: 220px; }}
  td.col-src .src-name {{ color: {L['text_muted']}; font-size: 11px; }}
  td.col-url {{ width: 320px; }}
  td.col-url .url-link {{
    color: {L['accent_ai']}; text-decoration: none;
    display: inline-block; max-width: 300px;
    overflow: hidden; text-overflow: ellipsis; white-space: nowrap;
    vertical-align: bottom;
  }}
  td.col-url .url-link:hover {{ text-decoration: underline; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover td {{ background: {L['bg_hover']}; }}
  .empty {{
    padding: 60px 20px; text-align: center; color: {L['text_muted']};
    font-size: 13px;
  }}
  /* ── Footer ── */
  .footer {{
    text-align: center; margin-top: 24px;
    font-size: 11px; color: {L['text_muted']};
  }}
  .footer a {{ color: {L['accent_ai']}; text-decoration: none; }}
  .footer a:hover {{ text-decoration: underline; }}
  @media (max-width: 720px) {{
    .header {{ flex-direction: column; align-items: flex-start; gap: 8px; }}
    .actions {{ flex-direction: column; align-items: stretch; }}
    .actions-summary {{ margin-left: 0; }}
    td.col-url, th:nth-child(4) {{ display: none; }}
  }}
</style>
</head>
<body>
<div class="container">
  <!-- Header -->
  <div class="header">
    <div class="header-left">
      <div class="logo">H</div>
      <div>
        <div class="header-title">热点地图 <span style="font-size:11px;color:{L['text_muted']};font-weight:400;">HOTSPOT MAP</span></div>
        <div class="header-subtitle">数据导出 · {now}</div>
      </div>
    </div>
    <div class="header-meta">v1.2.0</div>
  </div>

  <!-- Filter form (GET to /api/export = refresh preview) -->
  <form id="filter-form" method="get" action="/api/export" class="card">
    <div class="card-title">筛选条件</div>

    <div class="field-row">
      <span class="field-label">时间范围</span>
      <label class="opt">
        <input type="radio" name="time_range" value="24h" {"checked" if tr_checked["24h"] else ""}>
        <span class="opt-label">24 小时</span>
      </label>
      <label class="opt">
        <input type="radio" name="time_range" value="3d" {"checked" if tr_checked["3d"] else ""}>
        <span class="opt-label">3 天</span>
      </label>
      <label class="opt">
        <input type="radio" name="time_range" value="7d" {"checked" if tr_checked["7d"] else ""}>
        <span class="opt-label">7 天</span>
      </label>
    </div>

    <div class="field-row">
      <span class="field-label">内容类型</span>
      <label class="opt">
        <input type="checkbox" name="types" value="hotspot" {"checked" if type_checked["hotspot"] else ""}>
        <span class="opt-label">资讯</span>
        <span class="opt-hint">(科技/AI · 安全 · 金融 · 创业 · GitHub)</span>
      </label>
      <label class="opt">
        <input type="checkbox" name="types" value="bid" {"checked" if type_checked["bid"] else ""}>
        <span class="opt-label">标讯</span>
        <span class="opt-hint">(招标)</span>
      </label>
    </div>

    <div class="actions">
      <button type="submit" class="btn btn-ghost">更新预览</button>
      <button type="submit" class="btn btn-primary" formaction="/api/export/download">↓ 导出 XLSX</button>
      <div class="actions-summary">
        当前: <b>{tr_label}</b> · <b>{types_label}</b> · 共 <b>{full_count}</b> 条
      </div>
    </div>
  </form>

  <!-- Preview table -->
  <div class="preview">
    <div class="preview-head">
      <div class="preview-title">数据预览</div>
      <div class="preview-hint">{overflow_note}</div>
    </div>
    {f'''<table>
      <thead>
        <tr>
          <th>#</th>
          <th>标题</th>
          <th>来源 / 分类</th>
          <th>原文链接</th>
        </tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>''' if items else '<div class="empty">当前筛选条件下没有数据, 请调整时间范围或勾选内容类型</div>'}
  </div>

  <div class="footer">
    热点地图 &middot; 数据来源: 安全客 / Krebs / PortSwigger / SANS ISC / FreeBuf / 奇安信 / AVD / CNNVD / CNVD / 新浪财经 / 东方财富 / Hacker News / aihot.virxact.com / GitHub Trending / 中国政府采购网<br>
    <a href="/">← 返回首页</a>
  </div>
</div>

<script>
  // 改动筛选条件自动提交 (form GET → 刷新当前页, 保留 query string 同步)
  document.querySelectorAll('#filter-form input').forEach(el => {{
    el.addEventListener('change', () => {{
      // checkbox 至少保留一个 (浏览器 default 行为就是允许 0 选中, 但 UX 上至少要 1 个)
      // 这里不强制, 让用户能体验 "0 选中 → 空表"
      document.getElementById('filter-form').submit();
    }});
  }});
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# XLSX 构建
# ---------------------------------------------------------------------------
def build_xlsx(items: list) -> bytes:
    """构造 XLSX 文件 (3 列: 标题 / 来源 / 原文链接)。

    与 /api/favorites/export 保持一致的样式 (Microsoft YaHei 表头 + 蓝色超链接)。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "热点清单"

    # 表头样式
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = ["标题", "来源", "原文链接"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    body_font = Font(name="Microsoft YaHei", size=10)
    link_font = Font(name="Microsoft YaHei", size=10, color="0563C1", underline="single")
    for idx, it in enumerate(items, start=2):
        a = ws.cell(row=idx, column=1, value=it.title or "")
        a.font = body_font
        a.alignment = left
        b = ws.cell(row=idx, column=2, value=it.source or "")
        b.font = body_font
        b.alignment = left
        c = ws.cell(row=idx, column=3, value=str(it.url))
        c.hyperlink = str(it.url)
        c.font = link_font
        c.alignment = left

    # 列宽 (中文字符 1 字 ≈ 2 ascii 宽)
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 60

    # 冻结表头
    ws.freeze_panes = "A2"
    # 行高 (避免中文标题被截断)
    ws.row_dimensions[1].height = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_filename(time_range: TimeRange, types: set[str]) -> str:
    """生成下载文件名: hotspots-YYYY-MM-DD-{time_range}-{types}.xlsx"""
    date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    types_token = "-".join(sorted(types)) if types else "empty"
    return f"hotspots-{date}-{time_range.value}-{types_token}.xlsx"


# ---------------------------------------------------------------------------
# ETag + 缓存 (旧版默认页专用)
# ---------------------------------------------------------------------------
def _compute_etag(html: str) -> str:
    return hashlib.md5(html.encode("utf-8")).hexdigest()


def _atomic_write_text(path: str, content: str) -> None:
    """原子写入：先写 tmp，再 rename 替换。"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        dir=os.path.dirname(path), prefix=".export_", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(content)
        os.replace(tmp, path)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def rebuild_export_cache() -> str:
    """预生成：写 ``export_cache.html`` + ``export_cache.etag``。"""
    html = build_html()
    etag = _compute_etag(html)
    _atomic_write_text(_CACHE_HTML, html)
    _atomic_write_text(_CACHE_ETAG, etag)
    return etag


def get_cached_export() -> Tuple[Optional[str], Optional[str]]:
    """读 ``(html, etag)``；不存在返回 ``(None, None)``。"""
    if not _CACHE_HTML.exists() or not _CACHE_ETAG.exists():
        return None, None
    try:
        html = _CACHE_HTML.read_text(encoding="utf-8")
        etag = _CACHE_ETAG.read_text(encoding="utf-8").strip()
    except OSError:
        return None, None
    return html, etag


def get_or_build_html() -> Tuple[str, str]:
    """返回 ``(html, etag)``。缓存命中 → 直接返回；否则实时构建。"""
    html, etag = get_cached_export()
    if html is not None and etag:
        return html, etag
    html = build_html()
    etag = _compute_etag(html)
    return html, etag


def get_cached_etag() -> Optional[str]:
    """只读 etag（ETag 304 流程用）。"""
    if not _CACHE_ETAG.exists():
        return None
    try:
        return _CACHE_ETAG.read_text(encoding="utf-8").strip()
    except OSError:
        return None


__all__ = [
    "build_html",
    "build_export_html",
    "build_xlsx",
    "xlsx_filename",
    "parse_export_filters",
    "fetch_export_items",
    "rebuild_export_cache",
    "get_cached_export",
    "get_cached_etag",
    "get_or_build_html",
]
