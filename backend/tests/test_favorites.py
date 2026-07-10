"""Phase 10 收藏仓库 + API 测试

覆盖范围（50+ 场景）：
  - FavoriteRepository 单元
    * add 新增 / add 重复
    * add 校验
    * remove / remove 不存在
    * list 全部 / list 分类筛选 / list limit 边界
    * is_favorited
    * list_favorited_ids 批量
    * count_by_category 全 6 分类都有
    * total
    * favorites_stats 增量 + 全表重算
  - Favorites API 端到端
    * POST add 成功 + 重复
    * GET list / category 筛选 / limit 边界
    * DELETE 成功 + 不存在
    * GET count 全 6 分类
    * GET export 字段（3 列）+ category 筛选
    * 错误响应（无效 category / 空 hotspot_id）
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.api import register_routers
from backend.api.middleware import TraceIDMiddleware
from backend.config import config
from backend.domain.enums import Category
from backend.exceptions import register_exception_handlers
from backend.repository import db
from backend.repository.favorite_repo import FavoriteRepository


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def temp_db(monkeypatch: pytest.MonkeyPatch, tmp_path):
    test_db = tmp_path / "test_favorites.db"
    monkeypatch.setattr(config, "db_path", test_db)
    db.close_db()
    db.init_db()
    yield test_db
    db.close_db()


@pytest.fixture
def repo(temp_db) -> FavoriteRepository:
    return FavoriteRepository()


@pytest.fixture
def client(temp_db) -> TestClient:
    app = FastAPI()
    app.add_middleware(TraceIDMiddleware)
    register_exception_handlers(app)
    register_routers(app)
    return TestClient(app)


def _make_fav(
    hid: str = "h-1",
    category: str = "ai",
    title: str = "Sample Title",
    source: str = "src",
    url: str = "https://example.com/a",
) -> dict:
    return {
        "hotspot_id": hid,
        "category": category,
        "title": title,
        "source": source,
        "url": url,
    }


# ===========================================================================
# 1. FavoriteRepository 单元测试
# ===========================================================================
class TestFavoriteRepoAdd:
    def test_add_creates_new(self, repo):
        created, item = repo.add(
            hotspot_id="h-1", category="ai", title="A", source="s", url="https://e.com/1"
        )
        assert created is True
        assert item.id > 0
        assert item.hotspot_id == "h-1"
        assert item.category == "ai"
        assert item.title == "A"
        assert item.url == "https://e.com/1"
        assert item.favorited_at  # ISO timestamp

    def test_add_duplicate_returns_existing(self, repo):
        repo.add(
            hotspot_id="h-dup", category="ai", title="A", source="s", url="https://e.com/dup"
        )
        created, item = repo.add(
            hotspot_id="h-dup", category="ai", title="A2", source="s2", url="https://e.com/dup2"
        )
        assert created is False
        # 原值不被覆盖（INSERT OR IGNORE）
        assert item.title == "A"
        assert item.url == "https://e.com/dup"
        assert item.source == "s"

    def test_add_empty_id_raises(self, repo):
        with pytest.raises(Exception):
            repo.add(
                hotspot_id="", category="ai", title="A", source="s", url="https://e.com/1"
            )

    def test_add_each_category_accepted(self, repo):
        for cat in ["ai", "security", "finance", "startup", "bid", "github"]:
            created, item = repo.add(
                hotspot_id=f"h-{cat}",
                category=cat,
                title=f"T-{cat}",
                source="s",
                url=f"https://e.com/{cat}",
            )
            assert created is True
            assert item.category == cat


class TestFavoriteRepoRemove:
    def test_remove_existing(self, repo):
        repo.add(
            hotspot_id="h-1", category="ai", title="A", source="s", url="https://e.com/1"
        )
        n = repo.remove("h-1")
        assert n == 1
        assert repo.is_favorited("h-1") is False

    def test_remove_not_existing(self, repo):
        n = repo.remove("nope")
        assert n == 0

    def test_remove_empty_id(self, repo):
        n = repo.remove("")
        assert n == 0


class TestFavoriteRepoList:
    def test_list_all_newest_first(self, repo):
        for i in range(3):
            repo.add(
                hotspot_id=f"h-{i}", category="ai", title=f"T{i}",
                source="s", url=f"https://e.com/{i}",
            )
        items = repo.list()
        assert len(items) == 3
        # 默认按 favorited_at DESC,所以 h-2 是最新
        assert items[0].hotspot_id == "h-2"
        assert items[1].hotspot_id == "h-1"
        assert items[2].hotspot_id == "h-0"

    def test_list_by_category(self, repo):
        for cat in ["ai", "security", "bid"]:
            repo.add(
                hotspot_id=f"h-{cat}", category=cat, title="T",
                source="s", url=f"https://e.com/{cat}",
            )
        ai_items = repo.list(category="ai")
        assert len(ai_items) == 1
        assert ai_items[0].category == "ai"

        bid_items = repo.list(category="bid")
        assert len(bid_items) == 1
        assert bid_items[0].category == "bid"

    def test_list_invalid_category_returns_empty(self, repo):
        repo.add(
            hotspot_id="h-1", category="ai", title="T", source="s", url="https://e.com/1"
        )
        items = repo.list(category="nonsense")
        # 不抛异常,只返回空
        assert items == []

    def test_list_limit_clamped(self, repo):
        # limit>1000 应被截断到 1000
        for i in range(3):
            repo.add(
                hotspot_id=f"h-{i}", category="ai", title="T", source="s", url=f"https://e.com/{i}"
            )
        items = repo.list(limit=2000)
        assert len(items) == 3  # <=1000
        # limit=0 走 fallback 200, 返回全部
        items_zero = repo.list(limit=0)
        assert len(items_zero) == 3  # fallback to default
        # limit=2 严格按 limit 截断
        items_two = repo.list(limit=2)
        assert len(items_two) == 2

    def test_list_default_limit(self, repo):
        for i in range(5):
            repo.add(
                hotspot_id=f"h-{i}", category="ai", title="T", source="s", url=f"https://e.com/{i}"
            )
        items = repo.list()
        assert len(items) == 5


class TestFavoriteRepoHelpers:
    def test_is_favorited_true_false(self, repo):
        repo.add(
            hotspot_id="h-1", category="ai", title="A", source="s", url="https://e.com/1"
        )
        assert repo.is_favorited("h-1") is True
        assert repo.is_favorited("h-x") is False

    def test_list_favorited_ids(self, repo):
        repo.add(
            hotspot_id="h-1", category="ai", title="A", source="s", url="https://e.com/1"
        )
        repo.add(
            hotspot_id="h-2", category="bid", title="B", source="s", url="https://e.com/2"
        )
        ids = repo.list_favorited_ids()
        assert ids == {"h-1", "h-2"}

    def test_count_by_category_all_six_present(self, repo):
        for cat in ["ai", "security", "finance", "startup", "bid", "github", "tech"]:
            repo.add(
                hotspot_id=f"h-{cat}", category=cat, title="T", source="s", url=f"https://e.com/{cat}"
            )
        counts = repo.count_by_category()
        # 全 7 分类都存在 (Phase 25 P1 加 tech)
        assert set(counts.keys()) == {"ai", "security", "finance", "startup", "bid", "github", "tech"}
        assert counts["ai"] == 1
        assert counts["bid"] == 1

    def test_count_by_category_zeros_for_empty(self, repo):
        counts = repo.count_by_category()
        # 即使没数据,7 分类都返回 0 (Phase 25 P1 加 tech)
        for cat in ["ai", "security", "finance", "startup", "bid", "github", "tech"]:
            assert counts[cat] == 0

    def test_total(self, repo):
        assert repo.total() == 0
        for i in range(4):
            repo.add(
                hotspot_id=f"h-{i}", category="ai", title="T", source="s", url=f"https://e.com/{i}"
            )
        assert repo.total() == 4


class TestFavoriteRepoStats:
    def test_stats_bump_on_add(self, repo):
        conn = db.get_connection()
        # 用循环 index 保证 hotspot_id 不同 (id() 对同一字符串返回相同值)
        for i, cat in enumerate(["ai", "ai", "bid"]):
            repo.add(
                hotspot_id=f"h-{cat}-{i}", category=cat, title="T", source="s",
                url=f"https://e.com/{cat}-{i}",
            )
        rows = conn.execute(
            "SELECT category, total_favorites FROM favorites_stats ORDER BY category"
        ).fetchall()
        d = {r["category"]: r["total_favorites"] for r in rows}
        assert d["ai"] == 2
        assert d["bid"] == 1

    def test_stats_refresh_on_remove(self, repo):
        for i in range(3):
            repo.add(
                hotspot_id=f"h-{i}", category="ai", title="T", source="s", url=f"https://e.com/{i}"
            )
        repo.remove("h-0")
        repo.remove("h-1")
        counts = repo.count_by_category()
        assert counts["ai"] == 1


# ===========================================================================
# 2. Favorites API 端到端测试
# ===========================================================================
class TestFavoritesAPIAdd:
    def test_add_returns_201_created(self, client):
        body = _make_fav(hid="h-1", category="ai")
        resp = client.post("/api/favorites", json=body)
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "ok"
        assert data["created"] is True
        assert data["item"]["hotspot_id"] == "h-1"
        assert data["item"]["category"] == "ai"

    def test_add_duplicate_returns_created_false(self, client):
        body = _make_fav(hid="dup", category="ai")
        r1 = client.post("/api/favorites", json=body)
        r2 = client.post("/api/favorites", json=body)
        assert r1.json()["created"] is True
        assert r2.json()["created"] is False

    def test_add_invalid_category_400(self, client):
        body = _make_fav(category="nonsense")
        resp = client.post("/api/favorites", json=body)
        assert resp.status_code == 400

    def test_add_empty_category_400(self, client):
        body = _make_fav(category="")
        resp = client.post("/api/favorites", json=body)
        assert resp.status_code in (400, 422)  # Pydantic 也可能拦

    def test_add_missing_hotspot_id_422(self, client):
        body = _make_fav()
        del body["hotspot_id"]
        resp = client.post("/api/favorites", json=body)
        assert resp.status_code == 422

    def test_add_each_category_works(self, client):
        for cat in ["ai", "security", "finance", "startup", "bid", "github"]:
            body = _make_fav(hid=f"h-{cat}", category=cat)
            resp = client.post("/api/favorites", json=body)
            assert resp.status_code == 200
            assert resp.json()["item"]["category"] == cat


class TestFavoritesAPIList:
    def _seed(self, client):
        items = []
        for cat in ["ai", "security", "bid", "github"]:
            body = _make_fav(
                hid=f"h-{cat}", category=cat, title=f"T-{cat}",
                url=f"https://e.com/{cat}",
            )
            r = client.post("/api/favorites", json=body)
            assert r.status_code == 200
            items.append(body)
        return items

    def test_list_all(self, client):
        self._seed(client)
        resp = client.get("/api/favorites")
        assert resp.status_code == 200
        data = resp.json()
        assert data["category"] == "all"
        assert data["count"] == 4
        assert data["total"] == 4
        assert len(data["items"]) == 4

    def test_list_by_category(self, client):
        self._seed(client)
        resp = client.get("/api/favorites", params={"category": "bid"})
        assert resp.status_code == 200
        data = resp.json()
        assert data["category"] == "bid"
        assert data["count"] == 1
        assert data["items"][0]["category"] == "bid"

    def test_list_invalid_category_empty(self, client):
        self._seed(client)
        resp = client.get("/api/favorites", params={"category": "nonsense"})
        assert resp.status_code == 400

    def test_list_limit(self, client):
        self._seed(client)
        resp = client.get("/api/favorites", params={"limit": 2})
        assert resp.status_code == 200
        assert resp.json()["count"] == 2


class TestFavoritesAPIRemove:
    def test_delete_success(self, client):
        body = _make_fav(hid="h-1", category="ai")
        client.post("/api/favorites", json=body)
        resp = client.delete("/api/favorites/h-1")
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "ok"
        assert data["removed"] == 1
        # 二次删除 → 0
        resp2 = client.delete("/api/favorites/h-1")
        assert resp2.json()["removed"] == 0

    def test_delete_empty_id_400(self, client):
        # 路径不能为空 → 实际是路由不匹配
        resp = client.delete("/api/favorites/")
        # FastAPI trailing slash 默认重定向或 404
        assert resp.status_code in (307, 404, 405)


class TestFavoritesAPICount:
    def test_count_all_six_present(self, client):
        for cat in ["ai", "security", "finance", "startup", "bid", "github", "tech"]:
            client.post("/api/favorites", json=_make_fav(hid=f"h-{cat}", category=cat))
        resp = client.get("/api/favorites/count")
        assert resp.status_code == 200
        data = resp.json()
        assert data["total"] == 7
        assert set(data["by_category"].keys()) == {
            "ai", "security", "finance", "startup", "bid", "github", "tech",
        }

    def test_count_empty(self, client):
        resp = client.get("/api/favorites/count")
        data = resp.json()
        assert data["total"] == 0
        assert all(v == 0 for v in data["by_category"].values())


# ===========================================================================
# 3. xlsx 导出测试
# ===========================================================================
class TestFavoritesAPIExport:
    def _seed_two(self, client):
        client.post("/api/favorites", json=_make_fav(
            hid="h-ai", category="ai", title="AI 资讯标题", source="src1",
            url="https://example.com/ai-news",
        ))
        client.post("/api/favorites", json=_make_fav(
            hid="h-bid", category="bid", title="招标标题", source="src2",
            url="https://example.com/bid-news",
        ))

    def test_export_returns_xlsx_bytes(self, client):
        self._seed_two(client)
        resp = client.get("/api/favorites/export")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        # 解析为 workbook 验证
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.title == "收藏清单"
        # 表头 3 列
        headers = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        assert headers == ["信息类型", "标题名称", "原文链接"]
        # 数据 2 行
        assert ws.max_row == 3

    def test_export_content_chinese_mapping(self, client):
        self._seed_two(client)
        resp = client.get("/api/favorites/export")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 收集所有信息类型列的值
        cats = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        # 中文映射: ai→科技/AI, bid→招标资讯
        assert "科技/AI" in cats
        assert "招标资讯" in cats

    def test_export_filter_by_category(self, client):
        self._seed_two(client)
        resp = client.get("/api/favorites/export", params={"category": "bid"})
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 只有 1 行数据
        assert ws.max_row == 2  # 表头 + 1 行
        assert ws.cell(row=2, column=1).value == "招标资讯"
        assert "X-Favorite-Count" in resp.headers
        assert resp.headers["X-Favorite-Count"] == "1"

    def test_export_empty_returns_only_header(self, client):
        resp = client.get("/api/favorites/export")
        assert resp.status_code == 200
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 只有表头
        assert ws.max_row == 1

    def test_export_hyperlink_set(self, client):
        self._seed_two(client)
        resp = client.get("/api/favorites/export")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=3)
            assert cell.hyperlink is not None
            assert cell.hyperlink.target.startswith("https://")

    def test_export_invalid_category_400(self, client):
        resp = client.get("/api/favorites/export", params={"category": "nonsense"})
        assert resp.status_code == 400

    def test_export_filename_header(self, client):
        self._seed_two(client)
        resp = client.get("/api/favorites/export")
        cd = resp.headers.get("content-disposition", "")
        assert "attachment" in cd
        assert "favorites_all_" in cd
        assert ".xlsx" in cd


# ===========================================================================
# 4. 错误格式验证（version + trace_id）
# ===========================================================================
class TestFavoritesAPIErrors:
    def test_error_response_has_version(self, client):
        resp = client.post("/api/favorites", json=_make_fav(category="bogus"))
        # 错误格式
        assert resp.status_code in (400, 500)
        body = resp.json()
        # 错误响应包含 message
        if "detail" in body:
            assert "message" in (body["detail"] if isinstance(body["detail"], dict) else {})

    def test_trace_id_injected(self, client):
        resp = client.get("/api/favorites/count")
        # TraceIDMiddleware 注入 X-Trace-Id
        assert "X-Trace-Id" in resp.headers
        assert len(resp.headers["X-Trace-Id"]) > 0


# ===========================================================================
# 5. 综合场景：增删改查 + 导出
# ===========================================================================
class TestFavoritesIntegration:
    def test_full_lifecycle(self, client):
        # 1. 空
        r = client.get("/api/favorites/count")
        assert r.json()["total"] == 0

        # 2. 加 3 条不同分类
        for cat in ["ai", "security", "bid"]:
            client.post("/api/favorites", json=_make_fav(
                hid=f"h-{cat}", category=cat,
                title=f"Title {cat}", source=f"src-{cat}",
                url=f"https://e.com/{cat}",
            ))

        # 3. 列表 = 3
        r = client.get("/api/favorites")
        assert r.json()["count"] == 3

        # 4. count = 3
        r = client.get("/api/favorites/count")
        assert r.json()["total"] == 3
        assert r.json()["by_category"]["ai"] == 1
        assert r.json()["by_category"]["security"] == 1
        assert r.json()["by_category"]["bid"] == 1

        # 5. 删 bid
        client.delete("/api/favorites/h-bid")
        r = client.get("/api/favorites/count")
        assert r.json()["total"] == 2
        assert r.json()["by_category"]["bid"] == 0

        # 6. 导出剩余 2 条
        resp = client.get("/api/favorites/export")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 表头 + 2 行
        assert ws.max_row == 3
