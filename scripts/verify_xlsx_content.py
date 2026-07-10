"""验证 test_favorites_export.xlsx 内容"""
import io
from openpyxl import load_workbook

with open("test_favorites_export.xlsx", "rb") as f:
    wb = load_workbook(io.BytesIO(f.read()))
ws = wb.active
print(f"Sheet: {ws.title}")
print(f"Rows: {ws.max_row} (1 表头 + {ws.max_row - 1} 数据)")
print(f"Cols: {ws.max_column}")
print()
print(f"Column widths: A={ws.column_dimensions['A'].width} B={ws.column_dimensions['B'].width} C={ws.column_dimensions['C'].width}")
print(f"Frozen: {ws.freeze_panes}")
print()
print("=== Headers ===")
for c in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=c)
    print(f"  col {c}: '{cell.value}' (bold={cell.font.bold}, fill={cell.fill.start_color.rgb if cell.fill.start_color else 'none'})")
print()
print("=== Data ===")
for r in range(2, ws.max_row + 1):
    cat = ws.cell(row=r, column=1).value
    title = ws.cell(row=r, column=2).value
    url_cell = ws.cell(row=r, column=3)
    url = url_cell.value
    has_link = url_cell.hyperlink is not None
    print(f"  Row {r}: {cat} | {title} | {url} (hyperlink={has_link})")
